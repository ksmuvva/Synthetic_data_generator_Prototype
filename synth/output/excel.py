"""
Enhanced Excel output generator using openpyxl.

Generates Excel files with formatting, multiple sheets, and styles.
"""

from pathlib import Path
from typing import Optional
import pandas as pd
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from synth.output.base import OutputGenerator, GeneratorRegistry
from synth.patterns.schema import Schema


class ExcelGenerator(OutputGenerator):
    """Generate Excel files with enhanced formatting."""

    def __init__(self):
        """Initialize Excel generator."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel generation. "
                "Install it with: pip install openpyxl"
            )

    def generate(
        self,
        data: pd.DataFrame,
        output_path: Path,
        schema: Optional[Schema] = None,
        sheet_name: str = "Synthetic Data",
        include_metadata: bool = True,
        freeze_header: bool = True,
        auto_filter: bool = True,
        **kwargs
    ) -> Path:
        """
        Generate Excel file with formatting.

        Args:
            data: DataFrame to write
            output_path: Where to save the Excel file
            schema: Optional schema for metadata
            sheet_name: Name of the data sheet
            include_metadata: Include metadata sheet
            freeze_header: Freeze the header row
            auto_filter: Enable auto-filter on header row
            **kwargs: Additional options

        Returns:
            Path to the generated Excel file
        """
        # Ensure output path has .xlsx extension
        output_path = Path(output_path)
        if output_path.suffix not in (".xlsx", ".xls"):
            output_path = output_path.with_suffix(".xlsx")

        # Create parent directory if needed
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Add metadata sheet if schema provided
        if schema and include_metadata:
            self._add_metadata_sheet(wb, schema, data)

        # Add data sheet
        self._add_data_sheet(wb, data, sheet_name, schema, freeze_header, auto_filter)

        # Set active sheet to data
        wb.active = wb[sheet_name]

        # Save workbook
        wb.save(output_path)

        return output_path

    def _add_metadata_sheet(
        self,
        wb: Workbook,
        schema: Schema,
        data: pd.DataFrame
    ) -> None:
        """Add metadata sheet with schema information."""
        ws = wb.create_sheet("Metadata")

        # Title
        ws["A1"] = "Synthetic Data Metadata"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:B1")

        # Generation info
        ws["A3"] = "Generated:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A3"].font = Font(bold=True)

        # Data info
        row = 5
        ws[f"A{row}"] = "Row Count:"
        ws[f"B{row}"] = schema.row_count
        ws[f"A{row}"].font = Font(bold=True)

        row += 1
        ws[f"A{row}"] = "Number of Fields:"
        ws[f"B{row}"] = len(schema.fields)
        ws[f"A{row}"].font = Font(bold=True)

        row += 2
        ws[f"A{row}"] = "Fields:"
        ws[f"A{row}"].font = Font(bold=True, size=12)

        # Field details
        row += 1
        headers = ["Field Name", "Type", "Unique", "Nullable", "Min", "Max"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092",
                end_color="366092",
                fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=True)

        # Field data
        for field in schema.fields:
            row += 1
            ws.cell(row=row, column=1, value=field.name)
            ws.cell(row=row, column=2, value=field.type.value)
            ws.cell(row=row, column=3, value="Yes" if field.unique else "No")
            ws.cell(row=row, column=4, value="No" if field.nullable else "Yes")

            # Add min/max for numeric fields
            if field.type.value in ("integer", "float"):
                ws.cell(row=row, column=5, value=field.min_value)
                ws.cell(row=row, column=6, value=field.max_value)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col[0].letter].width = adjusted_width

    def _add_data_sheet(
        self,
        wb: Workbook,
        data: pd.DataFrame,
        sheet_name: str,
        schema: Optional[Schema],
        freeze_header: bool,
        auto_filter: bool
    ) -> None:
        """Add data sheet with formatted table."""
        ws = wb.create_sheet(sheet_name)

        # Add DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Apply formatting
        self._apply_header_format(ws, len(data.columns))
        self._apply_column_widths(ws, data)

        if freeze_header:
            ws.freeze_panes = "A2"

        if auto_filter:
            ws.auto_filter.ref = ws.dimensions

    def _apply_header_format(self, ws, num_cols: int) -> None:
        """Apply formatting to header row."""
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

    def _apply_column_widths(self, ws, data: pd.DataFrame) -> None:
        """Auto-adjust column widths."""
        for col_idx, col_name in enumerate(data.columns, 1):
            # Get column letter
            col_letter = chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"

            # Calculate width based on content
            max_length = max(
                len(str(col_name)),
                *([len(str(val)) for val in data[col_name][:100]] if col_name in data else [])
            )

            # Set width (with padding)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    def supports_format(self, format_type: str) -> bool:
        """Check if format is supported."""
        return format_type.lower() in ("excel", "xlsx", "xls")


# Register the generator
if OPENPYXL_AVAILABLE:
    GeneratorRegistry.register("excel", ExcelGenerator)
    GeneratorRegistry.register("xlsx", ExcelGenerator)
